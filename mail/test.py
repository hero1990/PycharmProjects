import pymssql as pms
from openpyxl import load_workbook
import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib
import pandas as pd


def get_data(sql, file_path):

    conn = pms.connect(host='10.213.27.220', user='sa',
                       password='mssql', database='KoneSH', charset="utf8")
    cur = conn.cursor()
    cur.execute(sql)
    # 获取所需要的字段名称
    fields = cur.description
    # 获取所需要的数据
    data = cur.fetchall()
    # 关闭连接
    cur.close()
    # 返回所需的数据
    pd.DataFrame(data).to_excel(file_path, index=False)
    generate_excel(fields, file_path)


def generate_excel(field, file):
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name('Sheet1')
    sheet.title = '数据展示'
    # 将字段名称循环写入excel第一行，因为字段格式列表里包含列表，每个列表的第一元素才是字段名称
    for col in range(len(field)):
        # row代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始，这点于python不同要注意
        _ = sheet.cell(row=1, column=col+1, value=u'%s' % field[col][0])
        wb.save(file)


def get_yesterday():
    # 获取昨天日期的字符串格式的函数
    # 获取今天的日期
    today = datetime.date.today()
    # 获取一天的日期格式数据
    oneday = datetime.timedelta(days=1)
    # 昨天等于今天减去一天
    yesterday = today - oneday
    # 获取昨天日期的格式化字符串
    yesterdaystr = yesterday.strftime('%Y-%m-%d')
    # 返回昨天的字符串
    return yesterdaystr


def create_email(me, mailto_list, email_subject, email_text, file_path, file_name):
    msg = MIMEMultipart()
    # 将正文以text的形式插入邮件中
    msg.attach(MIMEText(email_text, 'plain', 'utf-8'))
    msg['Subject'] = Header(email_subject, 'utf-8')
    msg['From'] = me
    msg['To'] = ";".join(mailto_list)
    att1 = MIMEText(open(file_path, 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # #生成附件的名称
    att1["Content-Disposition"] = 'attachment; filename=' + file_name
    # #将附件内容插入邮件中
    msg.attach(att1)
    return msg


def main():
    mailto_list = ["jack-sj.chen@dbschenker.com"]
    mail_host = "10.209.129.122"
    mail_user = "Schenker_wms@dbschenker.com"
    mail_pwd = "hero19901001"
    mail_from = mail_user
    print(datetime.datetime.now())
    sql = "SELECT * from orders where orderkey = '0000048036' ;"
    # 得到昨天的日期
    yesterday_str = get_yesterday()
    # 文件名称
    file_name = 'user attribute' + yesterday_str + '.xlsx'
    # 文件路径
    file_path = 'C:/work/report/' + file_name
    get_data(sql, file_path)
    # 邮件标题
    email_subject = 'test' + yesterday_str
    # 邮件正文
    email_text = "Dear all,\n\t附件为每周数据，请查收！\n\nBI团队 "

    # 生成邮件
    try:
        s = smtplib.SMTP(mail_host,25)
        #s.connect(mail_host)  # 连接smtp服务器
        #s.login(mail_user, mail_pwd)  # 登陆服务器
        msg = create_email(mail_from, mailto_list, email_subject, email_text, file_path, file_name)
        s.sendmail(mail_from, mailto_list, msg.as_string())  # 发送邮件
        s.close()
        print('successfully')
    except Exception as e:
        print(str(e))


if __name__ == "__main__":
    main()
